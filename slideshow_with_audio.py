from pydub import AudioSegment  # Обязательна предварительная установка ffmpeg и указание пути к нему в Windows
from moviepy.editor import *
import os
import random
import re
from os import walk
import shutil
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook


# Если добавлялись новые фоны, то перед запуском данного скрипта надо запустить скрипт "edited_fons".
# Перед запуском программы запускаем скрипт "resize", делающий исходные картинка resize и crop и помещающий их в

# Прописываем пути к папкам и файлам:

path_name_ozv = r'D:\Cardsmaker\mp3_ozv'  # путь до файлов с озвучкой и музыкой
path_name_tabs = r'D:\Cardsmaker\clients\mediatemp'  # путь до файла с таблицей
path_name_images = r'D:\Cardsmaker\clients\mediatemp\target'  # путь до папок с картинками, из которых создадим видео открытку
path_name_video = r'D:\Cardsmaker\clients\mediatemp\target\video'  # путь до папки с готовым видео
path_name_fons = r'D:\Cardsmaker\fons'  # путь до папки с фонами
path_name_edited_fons = r'D:\Cardsmaker\edited_fons'  # путь до папки с фонами, подготовленными к наложению голоса
filename_tab = 'tab_ozv.xlsx'  # имя файла с таблицами

# Создаем списки. Получаем их из таблицы:

wb = load_workbook(os.path.join(path_name_tabs, filename_tab), read_only=False)
sh = wb['women_names']
wb.active = sh

women_names = []  # создаем список имен
for row in sh.iter_cols(min_row=1, max_row=100, min_col=4, max_col=4):  # из какой колонки получаем данные
    for cell in row:
        women_names.append(cell.value)

stih_id_names = []  # создаем список id:
for row in sh.iter_cols(min_row=1, max_row=100, min_col=1, max_col=1):  # из какой колонки получаем данные
    for cell in row:
        stih_id_names.append(cell.value)

nsk_names = []  # создаем список nsk со страницы men_names:
for row in sh.iter_cols(min_row=1, max_row=100, min_col=2, max_col=2):  # из какой колонки получаем данные
    for cell in row:
        nsk_names.append(cell.value)

stih_with_name = []
for row in sh.iter_cols(min_row=1, max_row=100, min_col=5, max_col=5):
    for cell in row:
        stih_with_name.append(cell.value)

# открытваем страницу др-2
sh2 = wb['др-1']
wb.active = sh2

stih_id_begin = []
for row in sh2.iter_cols(min_row=1, max_row=24, min_col=1, max_col=1):
    for cell in row:
        stih_id_begin.append(cell.value)

stih_id_middle_1 = []
for row in sh2.iter_cols(min_row=25, max_row=171, min_col=1, max_col=1):
    for cell in row:
        stih_id_middle_1.append(cell.value)

stih_id_middle_2 = []
for row in sh2.iter_cols(min_row=172, max_row=310, min_col=1, max_col=1):
    for cell in row:
        stih_id_middle_2.append(cell.value)

stih_id_end = []
for row in sh2.iter_cols(min_row=311, max_row=375, min_col=1, max_col=1):
    for cell in row:
        stih_id_end.append(cell.value)

#####
stih_begin = []
for row in sh2.iter_cols(min_row=1, max_row=24, min_col=5, max_col=5):
    for cell in row:
        stih_begin.append(cell.value)

stih_middle_1 = []
for row in sh2.iter_cols(min_row=25, max_row=171, min_col=5, max_col=5):
    for cell in row:
        stih_middle_1.append(cell.value)

stih_middle_2 = []
for row in sh2.iter_cols(min_row=172, max_row=310, min_col=5, max_col=5):
    for cell in row:
        stih_middle_2.append(cell.value)

stih_end = []
for row in sh2.iter_cols(min_row=311, max_row=375, min_col=5, max_col=5):
    for cell in row:
        stih_end.append(cell.value)

edited_fons = os.listdir(path_name_edited_fons)


for j in range(0, 1):  # задаем кол-во имен для обработки
    woman_name = women_names[j]  # выбор имени в списке
    print(woman_name)
    images = os.listdir(os.path.join(path_name_images, woman_name))  # список уже обработанных изображений в папке 'images'
    print(images)

    try:
        os.mkdir(os.path.join(path_name_video, women_names[j]))  # создаем папку с именем для видео
    except FileExistsError:
        pass  # чтобы продолжать, если папка уже создана ранее

    for k in range(0, 5):
        slideshow = random.sample(images, 4)

    # file_body, ext = os.path.splitext(image0)

        fon = random.choice(edited_fons)  # случайный выбор фона
        fon = AudioSegment.from_mp3(os.sep.join([path_name_edited_fons, fon]))

############### Формируем именную озвучку ###############

        if nsk_names[j] == 'Начало':
            first = stih_id_names[j]
            first_stih = stih_with_name[j]
        else:
                first = random.choice(stih_id_begin)  # случайный выбор из списка "Начало"
                first_stih = stih_begin[stih_id_begin.index(first)]

        if nsk_names[j] == 'Середина':
            second = stih_id_names[j]
            second_stih = stih_with_name[j]
        else:
            second = random.choice(stih_id_middle_1)
            second_stih = stih_middle_1[stih_id_middle_1.index(second)]

        if nsk_names[j] == 'Конец':
            fourth = stih_id_names[j]
            fourth_stih = stih_with_name[j]

        else:
            fourth = random.choice(stih_id_end)
            fourth_stih = stih_end[stih_id_end.index(fourth)]

        third = random.choice(stih_id_middle_2)
        third_stih = stih_middle_2[stih_id_middle_2.index(third)]

        a = str(first)+'.mp3'  # формируем названия файла, по нему мы отыщем на компе нужный файл.
        b = str(second)+'.mp3'
        c = str(third)+'.mp3'
        d = str(fourth)+'.mp3'

        stih1 = str(first_stih)
        stih2 = str(second_stih)
        stih3 = str(third_stih)
        stih4 = str(fourth_stih)
        stih = stih1 + '\n\n' + stih2 + '\n\n' + stih3 + '\n\n' + stih4
        print(stih)

        spase = AudioSegment.silent(duration=1000)  # Аудиосегмент длительностью 1 сек. без звука
        try:
            tr_1 = AudioSegment.from_mp3(os.sep.join([path_name_ozv, a]))
            tr_2 = AudioSegment.from_mp3(os.sep.join([path_name_ozv, b]))
            tr_3 = AudioSegment.from_mp3(os.sep.join([path_name_ozv, c]))
            tr_4 = AudioSegment.from_mp3(os.sep.join([path_name_ozv, d]))

        except FileNotFoundError:
            continue
        t1 = tr_1.duration_seconds + 3
        t2 = tr_2.duration_seconds + 1
        t3 = tr_3.duration_seconds + 1
        t4 = 60 - (t1 + t2 + t3)

        videoclip0 = ImageClip(os.path.join(path_name_images, woman_name, slideshow[0]))
        videoclip0.duration = t1

        videoclip1 = ImageClip(os.path.join(path_name_images, woman_name, slideshow[1]))
        videoclip1.duration = t2

        videoclip2 = ImageClip(os.path.join(path_name_images, woman_name, slideshow[2]))
        videoclip2.duration = t3

        videoclip3 = ImageClip(os.path.join(path_name_images, woman_name, slideshow[3]))
        videoclip3.duration = t4

        # file_body, ext = os.path.splitext(image)

        # описание для открытки:
        set_pos = wb['comb']
        wb.active = set_pos

        words_1 = []
        for row in set_pos.iter_cols(min_row=1, max_row=8, min_col=1, max_col=1):
            for cell in row:
                words_1.append(cell.value)

        words_2 = []
        for row in set_pos.iter_cols(min_row=1, max_row=10, min_col=2, max_col=2):
            for cell in row:
                words_2.append(cell.value)

        words_3 = []
        for row in set_pos.iter_cols(min_row=1, max_row=9, min_col=3, max_col=3):
            for cell in row:
                words_3.append(cell.value)

        words_4 = []
        for row in set_pos.iter_cols(min_row=20, max_row=23, min_col=4, max_col=4):
            for cell in row:
                words_4.append(cell.value)

        words_5 = []
        for row in set_pos.iter_cols(min_row=1, max_row=2, min_col=5, max_col=5):
            for cell in row:
                words_5.append(cell.value)

        words_6 = []
        for row in set_pos.iter_cols(min_row=1, max_row=2, min_col=6, max_col=6):
            for cell in row:
                words_6.append(cell.value)

        words_7 = []
        for row in set_pos.iter_cols(min_row=1, max_row=4, min_col=7, max_col=7):
            for cell in row:
                words_7.append(cell.value)

        words_8 = []
        for row in set_pos.iter_cols(min_row=1, max_row=5, min_col=8, max_col=8):
            for cell in row:
                words_8.append(cell.value)

        comb1 = str(random.choice(words_1))
        comb2 = str(random.choice(words_2))
        comb3 = str(random.choice(words_3))
        comb4 = str(random.choice(words_4))
        comb5 = str(random.choice(words_5))
        comb6 = str(random.choice(words_6))
        comb7 = str(random.choice(words_7))
        comb8 = str(random.choice(words_8))

        comb = comb1 + str(" ") + comb2 + str(" ") + comb3 + str(" ") + comb4 + str(" ") + comb5 + str(
            " ") + woman_name + \
                str('.') + str(" ") + comb6 + str(" ") + comb7 + str(" ") + comb8 + '\n\n' + stih

        # создается и сохраняется файл с четверостишием в открытке
        file = open(os.path.join(path_name_video, women_names[j], str(k) + '.txt'), "w")
        file.write(comb)
        file.close()

        # из отдельных треков и пауз собираем озвучку

        playlist = tr_1 + spase + tr_2 + spase + tr_3 + spase + tr_4
        played_together = fon.overlay(playlist, position=2000)  # Наложение голоса на музыку с задержкой 2 сек.
        played_together.export("voice.mp3", format="mp3")  # Сохранение в файл.

        audioclip = AudioFileClip("voice.mp3")

        videoclip = concatenate_videoclips([videoclip0, videoclip1, videoclip2, videoclip3])
        videoclip.duration = 60

        videoclip = videoclip.set_audio(audioclip)  # добавляем аудио
        videoclip.write_videofile(os.sep.join([path_name_video, woman_name, str(k) + '.mp4']), fps=24)  # сохраняем видео